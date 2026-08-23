from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.contrib import admin, messages
from django.contrib.admin.models import LogEntry
from django.core.exceptions import PermissionDenied, ValidationError as DjangoValidationError
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.http import urlencode
from django.db.models import Count, Q, Sum
from .forms import (
    AchievementForm, AdmissionInfoForm, AdmissionInquiryAdminForm, BatchForm, ClassForm,
    ClassScheduleForm, ContactMessageAdminForm, ExamForm, ExamResultForm, FacultyForm,
    GalleryForm, NoticeForm, StudentForm, StudentPaymentForm, SubjectForm, TeacherSalaryForm,
)
import openpyxl
import re

 

from .models import (
    Class,
    Batch,
    Faculty,
    Subject,
    Student,
    ClassSchedule,
    Attendance,
    Exam,
    ExamResult,
    StudentPayment,
    TeacherSalary,
    Achievement,
    Gallery,
    Notice,
    AdmissionInfo,
    AdmissionInquiry,
    ContactMessage,
)

# ---------------------------------------------------------
# Site branding
# ---------------------------------------------------------

admin.site.site_header = "Pi - π Academy Admin"
admin.site.site_title = "Pi Academy Admin"
admin.site.index_title = "Dashboard"


# ---------------------------------------------------------
# Inlines — manage related records from the parent's page
# ---------------------------------------------------------

class BatchInline(admin.TabularInline):
    model = Batch
    extra = 0
    fields = ("batch_name", "room", "start_time", "end_time")


class StudentPaymentInline(admin.TabularInline):
    model = StudentPayment
    extra = 0
    fields = ("payment_type", "amount", "payment_date", "payment_month", "remarks")


class AttendanceInline(admin.TabularInline):
    model = Attendance
    extra = 0
    fields = ("batch", "attendance_date", "status", "remarks")
    autocomplete_fields = ("batch",)


class ExamResultInline(admin.TabularInline):
    model = ExamResult
    extra = 0
    fields = ("exam", "subject", "marks", "grade")
    autocomplete_fields = ("exam", "subject")


class ClassScheduleInline(admin.TabularInline):
    model = ClassSchedule
    extra = 0
    fields = ("subject", "teacher", "day_of_week", "start_time", "end_time", "room")
    autocomplete_fields = ("subject", "teacher")


class TeacherSalaryInline(admin.TabularInline):
    model = TeacherSalary
    extra = 0
    fields = ("salary_month", "amount", "payment_date", "status", "remarks")


# ---------------------------------------------------------
# Model admins
# ---------------------------------------------------------

@admin.register(Class)
class ClassAdmin(admin.ModelAdmin):
    # search_fields is still required even though the default changelist
    # is bypassed below — other admins (Student, AdmissionInquiry, ...)
    # use autocomplete_fields=("class_obj",), and Django's autocomplete
    # endpoint looks up search_fields on THIS admin to serve results.
    search_fields = ("class_name",)

    # ------------------------------------------------------------------
    # Custom dark-themed pages, wired to the SAME URLs Django's default
    # admin would generate (admin:core_class_changelist / _add / _change
    # / _delete) — so every existing link into this admin keeps working.
    # Trade-off: this replaces the default list/add/change/delete views
    # entirely, so BatchInline (adding batches from this page) no longer
    # applies. Batches have their own dedicated page now, so that's fine.
    # ------------------------------------------------------------------

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        classes = (
            Class.objects.all()
            .prefetch_related("batch_set")
            .order_by("-academic_year", "class_name")
        )
        context = dict(
            self.admin_site.each_context(request),
            classes=classes,
            active_section="classes",
            active_page="class_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/class/list.html", context)

    def get_urls(self):
        # Adds a read-only "view" page alongside the existing add/change/
        # delete URLs, at admin/core/class/<id>/view/, reverse()-able as
        # admin:core_class_view.
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_class_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._class_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Class, pk=object_id)
        return self._class_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Class, pk=object_id)
        batches = instance.batch_set.all().order_by("batch_name")
        students = (
            Student.objects.filter(class_obj=instance)
            .select_related("batch")
            .order_by("full_name")
        )
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            batches=batches,
            students=students,
            active_section="classes",
            active_page="class_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/class/detail.html", context)

    def _class_form_view(self, request, instance):
        if request.method == "POST":
            form = ClassForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                self.message_user(
                    request,
                    f"Class {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_class_changelist")
        else:
            form = ClassForm(instance=instance)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="classes",
            active_page="class_edit" if instance else "class_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/class/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(Class, pk=object_id)
            obj.delete()
            self.message_user(request, "Class deleted.", level=messages.SUCCESS)
        return redirect("admin:core_class_changelist")


@admin.register(Batch)
class BatchAdmin(admin.ModelAdmin):
    # Required for autocomplete on other admins (e.g. Student, ClassSchedule
    # use autocomplete_fields including "batch") — see note on ClassAdmin.
    search_fields = ("batch_name", "room")

    # Same pattern as ClassAdmin: same admin URLs, custom dark templates.
    # Trade-off: ClassScheduleInline (adding schedules from this page) no
    # longer applies — manage those from Class Schedules in the admin
    # directly, which is unaffected by this change.

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        batches = Batch.objects.select_related("class_obj").order_by(
            "class_obj__class_name", "batch_name"
        )
        class_filter_id = request.GET.get("class")
        active_class = None
        if class_filter_id:
            batches = batches.filter(class_obj_id=class_filter_id)
            active_class = Class.objects.filter(pk=class_filter_id).first()
        context = dict(
            self.admin_site.each_context(request),
            batches=batches,
            active_class=active_class,
            active_section="batches",
            active_page="batch_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/batch/list.html", context)

    def get_urls(self):
        # Adds a read-only "view" page alongside the existing add/change/
        # delete URLs, at admin/core/batch/<id>/view/, reverse()-able as
        # admin:core_batch_view.
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_batch_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        lock_class = None
        class_id = request.GET.get("class")
        if class_id:
            lock_class = get_object_or_404(Class, pk=class_id)
        return self._batch_form_view(request, instance=None, lock_class=lock_class)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Batch, pk=object_id)
        return self._batch_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(
            Batch.objects.select_related("class_obj"), pk=object_id
        )
        students = (
            Student.objects.filter(batch=instance)
            .select_related("class_obj")
            .order_by("full_name")
        )
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            students=students,
            active_section="batches",
            active_page="batch_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/batch/detail.html", context)

    def _batch_form_view(self, request, instance, lock_class=None):
        if request.method == "POST":
            form = BatchForm(request.POST, instance=instance, lock_class=lock_class)
            if form.is_valid():
                form.save()
                self.message_user(
                    request,
                    f"Batch {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                if lock_class is not None:
                    return redirect(
                        reverse("admin:core_batch_changelist") + f"?class={lock_class.pk}"
                    )
                return redirect("admin:core_batch_changelist")
        else:
            form = BatchForm(instance=instance, lock_class=lock_class)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            locked_class=lock_class,
            active_section="batches",
            active_page="batch_edit" if instance else "batch_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/batch/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(Batch, pk=object_id)
            obj.delete()
            self.message_user(request, "Batch deleted.", level=messages.SUCCESS)
        return redirect("admin:core_batch_changelist")


@admin.register(Faculty)
class FacultyAdmin(admin.ModelAdmin):
    # Required for autocomplete on other admins (ClassScheduleInline,
    # TeacherSalaryAdmin use autocomplete_fields=("teacher",)) — see
    # note on ClassAdmin.
    search_fields = ("full_name", "phone", "email")

    # Same pattern as ClassAdmin/BatchAdmin: same admin URLs, custom
    # templates. Trade-off: ClassScheduleInline/TeacherSalaryInline no
    # longer show on this page — the detail page below lists both
    # instead.

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        faculty_list = Faculty.objects.all().order_by("full_name")
        context = dict(
            self.admin_site.each_context(request),
            faculty_list=faculty_list,
            active_section="faculty",
            active_page="faculty_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/faculty/list.html", context)

    def get_urls(self):
        # Adds a read-only "view" page alongside the existing add/change/
        # delete URLs, at admin/core/faculty/<id>/view/, reverse()-able
        # as admin:core_faculty_view.
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_faculty_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._faculty_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Faculty, pk=object_id)
        return self._faculty_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Faculty, pk=object_id)
        schedule = (
            ClassSchedule.objects.filter(teacher=instance)
            .select_related("batch", "batch__class_obj", "subject")
            .order_by("day_of_week", "start_time")
        )
        salaries = (
            TeacherSalary.objects.filter(teacher=instance)
            .order_by("-salary_month")
        )
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            schedule=schedule,
            salaries=salaries,
            active_section="faculty",
            active_page="faculty_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/faculty/detail.html", context)

    def _faculty_form_view(self, request, instance):
        if request.method == "POST":
            form = FacultyForm(request.POST, request.FILES, instance=instance)
            if form.is_valid():
                form.save()
                self.message_user(
                    request,
                    f"Faculty member {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_faculty_changelist")
        else:
            form = FacultyForm(instance=instance)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="faculty",
            active_page="faculty_edit" if instance else "faculty_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/faculty/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(Faculty, pk=object_id)
            obj.delete()
            self.message_user(request, "Faculty member deleted.", level=messages.SUCCESS)
        return redirect("admin:core_faculty_changelist")


@admin.register(Subject)
class SubjectAdmin(admin.ModelAdmin):
    # Required for autocomplete on other admins (ClassScheduleInline,
    # ExamResultAdmin use autocomplete_fields=("subject",)) — see note
    # on ClassAdmin.
    search_fields = ("subject_name", "subject_code")

    # Same pattern as ClassAdmin/FacultyAdmin: same admin URLs, custom
    # templates.

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        subjects = (
            Subject.objects.select_related("class_obj")
            .order_by("class_obj__class_name", "subject_name")
        )
        context = dict(
            self.admin_site.each_context(request),
            subjects=subjects,
            active_section="subjects",
            active_page="subject_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/subject/list.html", context)

    def get_urls(self):
        # Adds a read-only "view" page alongside the existing add/change/
        # delete URLs, at admin/core/subject/<id>/view/, reverse()-able
        # as admin:core_subject_view. Also adds the live subject-code
        # preview endpoint the Add Subject form calls.
        urls = super().get_urls()
        custom_urls = [
            path(
                "preview-code/",
                self.admin_site.admin_view(self.preview_code_view),
                name="core_subject_preview_code",
            ),
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_subject_view",
            ),
        ]
        return custom_urls + urls

    def preview_code_view(self, request):
        """
        AJAX endpoint the Add Subject form calls (via fetch) whenever
        the subject name or class changes, so staff see the code build
        before saving. Reuses _generate_subject_code directly so the
        preview can't drift from what actually gets saved — it's still
        just a preview, since a same-named subject saved a moment later
        by someone else could shift the collision-avoidance suffix.
        """
        if not self.has_add_permission(request):
            raise PermissionDenied
        subject_name = request.GET.get("subject_name", "").strip()
        class_id = request.GET.get("class_obj")
        if not subject_name or not class_id:
            return JsonResponse({"code": None})
        class_obj = Class.objects.filter(pk=class_id).first()
        if not class_obj:
            return JsonResponse({"code": None})
        return JsonResponse({"code": _generate_subject_code(subject_name, class_obj)})

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._subject_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Subject, pk=object_id)
        return self._subject_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Subject, pk=object_id)
        schedule = (
            ClassSchedule.objects.filter(subject=instance)
            .select_related("batch", "batch__class_obj", "teacher")
            .order_by("day_of_week", "start_time")
        )
        exams = (
            Exam.objects.filter(examresult__subject=instance)
            .distinct()
            .select_related("class_obj")
            .order_by("-exam_date")
        )
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            schedule=schedule,
            exams=exams,
            active_section="subjects",
            active_page="subject_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/subject/detail.html", context)

    def _subject_form_view(self, request, instance):
        if request.method == "POST":
            form = SubjectForm(request.POST, instance=instance)
            if form.is_valid():
                subject = form.save(commit=False)
                if not instance:
                    # Only generate a code for brand-new subjects — never
                    # regenerate one for an existing subject on edit.
                    subject.subject_code = _generate_subject_code(
                        subject.subject_name, subject.class_obj
                    )
                subject.save()
                self.message_user(
                    request,
                    f"Subject {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_subject_changelist")
        else:
            form = SubjectForm(instance=instance)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="subjects",
            active_page="subject_edit" if instance else "subject_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/subject/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(Subject, pk=object_id)
            obj.delete()
            self.message_user(request, "Subject deleted.", level=messages.SUCCESS)
        return redirect("admin:core_subject_changelist")


 


def _extract_number(text, default=1):
    """Pulls the first run of digits out of a string — 'Class 4' -> 4."""
    match = re.search(r"(\d+)", text or "")
    return int(match.group(1)) if match else default


def _subject_code_prefix(subject_name):
    """First 3 letters of the subject name, title-cased: 'Science' -> 'Sci'."""
    cleaned = re.sub(r"[^A-Za-z]", "", subject_name or "")
    if not cleaned:
        # Fallback for names with no usable Latin letters (e.g. pure
        # Bangla script) — still need something short and stable.
        return "Sub"
    return cleaned[:3].capitalize()


def _class_code_part(class_obj):
    """'Class Pronoy' -> 'classpronoy'; 'Class 6' -> 'class6'."""
    slug = re.sub(r"[^A-Za-z0-9]", "", class_obj.class_name or "").lower()
    return slug or f"class{class_obj.id}"


def _generate_subject_code(subject_name, class_obj):
    """
    Builds codes like Sci-class6 — first 3 letters of the subject name,
    then the class. Staff never type this in; see SubjectForm and
    SubjectAdmin._subject_form_view.
    """
    prefix = f"{_subject_code_prefix(subject_name)}-{_class_code_part(class_obj)}"
    code = prefix
    suffix = 1
    # Guards against collisions within the same class (e.g. "Science"
    # and "Sciences" both in Class 6 -> both start with "Sci-class6").
    while Subject.objects.filter(subject_code=code).exists():
        suffix += 1
        code = f"{prefix}{suffix}"
    return code


def _pagination_range(current, total, window=2):
    """
    Builds a page-number list with None as an ellipsis marker, e.g.
    [1, None, 4, 5, 6, None, 12] for page 5 of 12. Always keeps the
    first and last page visible, plus a window around the current page.
    """
    pages = []
    for p in range(1, total + 1):
        if p == 1 or p == total or (current - window <= p <= current + window):
            pages.append(p)
        elif pages and pages[-1] is not None:
            pages.append(None)
    return pages


def _generate_student_code(class_obj, batch):
    """
    Builds codes like PiC4B1001 — Pi Academy, Class 4, 1st batch of that
    class, 1st student in that class+batch. Staff never type this in.
    """
    class_num = _extract_number(class_obj.class_name, default=class_obj.id)

    if batch:
        sibling_ids = list(
            Batch.objects.filter(class_obj=class_obj).order_by("id").values_list("id", flat=True)
        )
        batch_num = sibling_ids.index(batch.id) + 1 if batch.id in sibling_ids else 1
    else:
        batch_num = 0  # no batch assigned yet

    prefix = f"PiC{class_num}B{batch_num}"
    seq = Student.objects.filter(student_code__startswith=prefix).count() + 1
    code = f"{prefix}{seq:03d}"

    # Guards against a stale count after deletions causing a collision.
    while Student.objects.filter(student_code=code).exists():
        seq += 1
        code = f"{prefix}{seq:03d}"
    return code


# Column headers for the bulk student enrollment template — order matters,
# since the upload view reads cells by position after checking these
# headers match. Keep this in sync with the template generator and the
# row parser below.
BULK_STUDENT_HEADERS = [
    "Full Name",
    "Class Name",
    "Academic Year",
    "Batch Name",
    "Gender",
    "Date of Birth (YYYY-MM-DD)",
    "Student Phone",
    "Father's Name",
    "Father's Phone",
    "Mother's Name",
    "Mother's Phone",
    "Address",
    "Admission Date (YYYY-MM-DD)",
    "Status",
]


def _parse_bulk_date(value):
    """Accepts an Excel-native date/datetime cell, or a YYYY-MM-DD /
    DD/MM/YYYY / DD-MM-YYYY string. Returns None if empty or unparsable."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


# Fixed leading columns for the exam result mark-sheet template — the
# rest of the header row is one column per subject in that exam's class,
# generated dynamically since it differs exam to exam.
BULK_EXAMRESULT_FIXED_HEADERS = ["Student ID", "Full Name"]


@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    # Still required: Attendance/ExamResult/StudentPayment admins use
    # autocomplete_fields=("student",), which relies on search_fields here.
    search_fields = (
        "student_code",
        "full_name",
        "student_phone",
        "father_name",
        "father_phone",
        "mother_name",
        "mother_phone",
    )

    # Same pattern as ClassAdmin/BatchAdmin: same admin URLs, custom light
    # templates. Trade-off: StudentPaymentInline / AttendanceInline /
    # ExamResultInline no longer show on this page — manage those from
    # their own sections in the sidebar instead.

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        students = Student.objects.select_related("class_obj", "batch").order_by("full_name")

        class_filter_id = request.GET.get("class")
        active_class = None
        if class_filter_id:
            students = students.filter(class_obj_id=class_filter_id)
            active_class = Class.objects.filter(pk=class_filter_id).first()

        query = request.GET.get("q", "").strip()
        if query:
            students = students.filter(
                Q(full_name__icontains=query)
                | Q(student_code__icontains=query)
                | Q(student_phone__icontains=query)
            )

        # Search/filter narrows the full table first — pagination only
        # applies to whatever that search matched, so a search always
        # finds a student regardless of which page they'd normally fall
        # on, and never needs "page 1" to be selected first.
        total_count = students.count()
        paginator = Paginator(students, 10)
        page_obj = paginator.get_page(request.GET.get("page"))

        preserved_params = request.GET.copy()
        preserved_params.pop("page", None)
        preserved_querystring = preserved_params.urlencode()

        context = dict(
            self.admin_site.each_context(request),
            students=page_obj,
            page_obj=page_obj,
            page_range=_pagination_range(page_obj.number, paginator.num_pages),
            preserved_querystring=preserved_querystring,
            total_count=total_count,
            classes=Class.objects.order_by("-academic_year", "class_name"),
            active_class=active_class,
            search_query=query,
            active_section="students",
            active_page="student_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/student/list.html", context)

    def get_urls(self):
        # Adds a read-only "view" page alongside the existing add/change/
        # delete URLs, at admin/core/student/<id>/view/, reverse()-able as
        # admin:core_student_view. Also adds the live student-code
        # preview endpoint and the bulk enrollment pages.
        urls = super().get_urls()
        custom_urls = [
            path(
                "bulk-upload/",
                self.admin_site.admin_view(self.bulk_upload_view),
                name="core_student_bulk_upload",
            ),
            path(
                "bulk-template/",
                self.admin_site.admin_view(self.bulk_template_view),
                name="core_student_bulk_template",
            ),
            path(
                "preview-code/",
                self.admin_site.admin_view(self.preview_code_view),
                name="core_student_preview_code",
            ),
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_student_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._student_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Student, pk=object_id)
        return self._student_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(
            Student.objects.select_related("class_obj", "batch"), pk=object_id
        )
        payments = StudentPayment.objects.filter(student=instance).order_by("-payment_date")
        total_paid = payments.aggregate(total=Sum("amount"))["total"] or 0
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            payments=payments,
            total_paid=total_paid,
            active_section="students",
            active_page="student_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/student/detail.html", context)

    def _student_form_view(self, request, instance):
        if request.method == "POST":
            form = StudentForm(request.POST, instance=instance)
            if form.is_valid():
                student = form.save(commit=False)
                if not instance:
                    # Only generate a code for brand-new students — never
                    # regenerate one for an existing student on edit.
                    student.student_code = _generate_student_code(
                        student.class_obj, student.batch
                    )
                student.save()
                self.message_user(
                    request,
                    f"Student {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_student_changelist")
        else:
            form = StudentForm(instance=instance)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="students",
            active_page="student_edit" if instance else "student_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/student/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(Student, pk=object_id)
            obj.delete()
            self.message_user(request, "Student deleted.", level=messages.SUCCESS)
        return redirect("admin:core_student_changelist")

    def preview_code_view(self, request):
        """
        AJAX endpoint the Add Student form calls (via fetch) whenever
        Class or Batch changes, so staff see what student_code WILL be
        before saving. Reuses _generate_student_code directly rather
        than duplicating the logic in JS, so the preview can never
        drift out of sync with what actually gets saved. It's still
        just a preview — the real code is generated again at save time,
        which is what actually reserves it.
        """
        if not self.has_add_permission(request):
            raise PermissionDenied
        class_id = request.GET.get("class_obj")
        if not class_id:
            return JsonResponse({"code": None})
        class_obj = Class.objects.filter(pk=class_id).first()
        if not class_obj:
            return JsonResponse({"code": None})
        batch_id = request.GET.get("batch")
        batch = Batch.objects.filter(pk=batch_id, class_obj=class_obj).first() if batch_id else None
        return JsonResponse({"code": _generate_student_code(class_obj, batch)})

    def bulk_template_view(self, request):
        if not self.has_add_permission(request):
            raise PermissionDenied
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Students"
        sheet.append(BULK_STUDENT_HEADERS)
        sheet.append([
            "Jane Doe", "Class 6", 2026, "Morning", "Female",
            "2014-05-12", "01700000000", "John Doe", "01700000001",
            "Mary Doe", "01700000002", "House 12, Road 4, Dhaka",
            "2026-01-10", "Active",
        ])
        for col_idx in range(1, len(BULK_STUDENT_HEADERS) + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="student_bulk_upload_template.xlsx"'
        return response

    def bulk_upload_view(self, request):
        if not self.has_add_permission(request):
            raise PermissionDenied

        results = None
        if request.method == "POST":
            excel_file = request.FILES.get("excel_file")
            if not excel_file:
                self.message_user(request, "Please choose a file to upload.", level=messages.ERROR)
            else:
                results = self._process_bulk_upload(excel_file)

        context = dict(
            self.admin_site.each_context(request),
            results=results,
            active_section="students",
            active_page="student_bulk_upload",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/student/bulk_upload.html", context)

    def _process_bulk_upload(self, excel_file):
        """
        Reads the uploaded workbook row by row, creating one Student per
        valid row. Processes sequentially (not bulk_create) so each
        student_code generation sees the previous row's save — this is
        what keeps sequence numbers correct when a file adds several
        students to the same class+batch at once.

        Returns a dict with created_count, errors (list of "Row N: ..."
        strings), and created (list of (name, code) for the summary).
        """
        try:
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception:
            return {"created_count": 0, "created": [], "errors": [
                "Couldn't read that file — make sure it's a valid .xlsx file, "
                "ideally the downloaded template with your rows filled in."
            ]}

        sheet = workbook.active
        header_row = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
        expected = [h.split(" (")[0] for h in BULK_STUDENT_HEADERS]
        got = [h.split(" (")[0] for h in header_row[:len(expected)]]
        if got != expected:
            return {"created_count": 0, "created": [], "errors": [
                "The column headers don't match the template. Please download "
                "the template below and use it as-is, filling in rows beneath it."
            ]}

        created = []
        errors = []
        status_values = {c[0].lower(): c[0] for c in Student._meta.get_field("status").choices}
        gender_values = {"male": "Male", "female": "Female", "other": "Other"}

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v in (None, "") for v in row):
                continue  # skip fully blank rows

            (full_name, class_name, academic_year, batch_name, gender,
             dob, phone, father_name, father_phone, mother_name,
             mother_phone, address, admission_date, status) = (
                list(row) + [None] * (len(BULK_STUDENT_HEADERS) - len(row))
            )[:len(BULK_STUDENT_HEADERS)]

            full_name = (str(full_name).strip() if full_name else "")
            if not full_name:
                errors.append(f"Row {row_num}: missing Full Name — skipped.")
                continue

            class_name = (str(class_name).strip() if class_name else "")
            if not class_name or not academic_year:
                errors.append(f"Row {row_num} ({full_name}): missing Class Name or Academic Year — skipped.")
                continue
            try:
                academic_year_int = int(academic_year)
            except (TypeError, ValueError):
                errors.append(f"Row {row_num} ({full_name}): Academic Year must be a number — skipped.")
                continue
            class_obj = Class.objects.filter(
                class_name__iexact=class_name, academic_year=academic_year_int
            ).first()
            if not class_obj:
                errors.append(
                    f"Row {row_num} ({full_name}): no class matching "
                    f"'{class_name}' ({academic_year_int}) — skipped."
                )
                continue

            batch = None
            batch_name = (str(batch_name).strip() if batch_name else "")
            if batch_name:
                batch = Batch.objects.filter(class_obj=class_obj, batch_name__iexact=batch_name).first()
                if not batch:
                    errors.append(
                        f"Row {row_num} ({full_name}): no batch matching "
                        f"'{batch_name}' in {class_obj.class_name} — skipped."
                    )
                    continue

            resolved_gender = gender_values.get(str(gender).strip().lower()) if gender else ""
            resolved_status = status_values.get(str(status).strip().lower()) if status else "Active"
            if status and not resolved_status:
                resolved_status = "Active"

            student = Student(
                class_obj=class_obj,
                batch=batch,
                full_name=full_name,
                gender=resolved_gender,
                date_of_birth=_parse_bulk_date(dob),
                student_phone=str(phone).strip() if phone else "",
                father_name=str(father_name).strip() if father_name else "",
                father_phone=str(father_phone).strip() if father_phone else "",
                mother_name=str(mother_name).strip() if mother_name else "",
                mother_phone=str(mother_phone).strip() if mother_phone else "",
                address=str(address).strip() if address else "",
                admission_date=_parse_bulk_date(admission_date),
                status=resolved_status,
            )
            try:
                student.full_clean(exclude=["student_code"])
            except DjangoValidationError as exc:
                errors.append(f"Row {row_num} ({full_name}): {'; '.join(exc.messages)} — skipped.")
                continue

            student.student_code = _generate_student_code(class_obj, batch)
            student.save()
            created.append((student.full_name, student.student_code))

        return {"created_count": len(created), "created": created, "errors": errors}

@admin.register(ClassSchedule)
class ClassScheduleAdmin(admin.ModelAdmin):
    # Required for autocomplete lookups (kept for parity with the other
    # admins in case a future inline references schedule entries).
    search_fields = ("subject__subject_name", "teacher__full_name", "batch__batch_name")

    # Same pattern as ClassAdmin/BatchAdmin/SubjectAdmin/StudentAdmin:
    # same admin URLs, custom templates.

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        schedule = (
            ClassSchedule.objects.select_related(
                "class_obj", "batch", "subject", "teacher"
            ).order_by(
                "class_obj__class_name", "batch__batch_name",
                "day_of_week", "start_time",
            )
        )
        class_filter_id = request.GET.get("class")
        active_class = None
        if class_filter_id:
            schedule = schedule.filter(class_obj_id=class_filter_id)
            active_class = Class.objects.filter(pk=class_filter_id).first()

        batch_filter_id = request.GET.get("batch")
        active_batch = None
        if batch_filter_id:
            schedule = schedule.filter(batch_id=batch_filter_id)
            active_batch = (
                Batch.objects.select_related("class_obj")
                .filter(pk=batch_filter_id)
                .first()
            )
        context = dict(
            self.admin_site.each_context(request),
            schedule=schedule,
            active_class=active_class,
            active_batch=active_batch,
            active_section="schedules",
            active_page="classschedule_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/classschedule/list.html", context)

    def get_urls(self):
        # Adds a read-only "view" page alongside the existing add/change/
        # delete URLs, at admin/core/classschedule/<id>/view/, reverse()-
        # able as admin:core_classschedule_view.
        urls = super().get_urls()
        custom_urls = [
            path(
                "class-options/",
                self.admin_site.admin_view(self.class_options_view),
                name="core_classschedule_class_options",
            ),
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_classschedule_view",
            ),
        ]
        return custom_urls + urls

    def class_options_view(self, request):
        """
        Returns the Subjects and Batches that belong to a given Class, as
        JSON. Powers the live filtering on the Add/Edit Class Schedule
        form so staff only see options that actually belong to the class
        they picked — mirrors the cross-check already enforced in
        ClassScheduleForm.clean().
        """
        if not self.has_view_permission(request):
            raise PermissionDenied
        class_id = request.GET.get("class_obj")
        class_obj = Class.objects.filter(pk=class_id).first() if class_id else None
        if not class_obj:
            return JsonResponse({"subjects": [], "batches": []})
        subjects = list(
            Subject.objects.filter(class_obj=class_obj)
            .order_by("subject_name")
            .values("id", "subject_name")
        )
        batches = list(
            Batch.objects.filter(class_obj=class_obj)
            .order_by("batch_name")
            .values("id", "batch_name")
        )
        return JsonResponse({
            "subjects": [{"id": str(s["id"]), "name": s["subject_name"]} for s in subjects],
            "batches": [{"id": str(b["id"]), "name": b["batch_name"]} for b in batches],
        })

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        lock_batch = None
        batch_id = request.GET.get("batch")
        if batch_id:
            lock_batch = Batch.objects.filter(pk=batch_id).first()
        lock_class = None
        if not lock_batch:
            class_id = request.GET.get("class")
            if class_id:
                lock_class = Class.objects.filter(pk=class_id).first()
        return self._classschedule_form_view(
            request, instance=None, lock_batch=lock_batch, lock_class=lock_class
        )

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(ClassSchedule, pk=object_id)
        return self._classschedule_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(
            ClassSchedule.objects.select_related(
                "class_obj", "batch", "subject", "teacher"
            ),
            pk=object_id,
        )
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            active_section="schedules",
            active_page="classschedule_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/classschedule/detail.html", context)

    def _classschedule_form_view(self, request, instance, lock_batch=None, lock_class=None):
        if request.method == "POST":
            form = ClassScheduleForm(
                request.POST, instance=instance, lock_batch=lock_batch, lock_class=lock_class
            )
            if form.is_valid():
                form.save()
                self.message_user(
                    request,
                    f"Class schedule entry {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                if lock_batch:
                    return redirect(
                        f"{reverse('admin:core_classschedule_changelist')}?batch={lock_batch.pk}"
                    )
                if lock_class:
                    return redirect(
                        f"{reverse('admin:core_classschedule_changelist')}?class={lock_class.pk}"
                    )
                return redirect("admin:core_classschedule_changelist")
        else:
            form = ClassScheduleForm(instance=instance, lock_batch=lock_batch, lock_class=lock_class)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="schedules",
            active_page="classschedule_edit" if instance else "classschedule_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/classschedule/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(ClassSchedule, pk=object_id)
            obj.delete()
            self.message_user(request, "Class schedule entry deleted.", level=messages.SUCCESS)
        return redirect("admin:core_classschedule_changelist")


@admin.register(Attendance)
class AttendanceAdmin(admin.ModelAdmin):
    list_display = (
        "student",
        "batch",
        "attendance_date",
        "status",
    )

    list_filter = (
        "attendance_date",
        "status",
        "batch",
    )

    search_fields = (
        "student__full_name",
        "student__student_code",
    )

    autocomplete_fields = ("student", "batch")


@admin.register(Exam)
class ExamAdmin(admin.ModelAdmin):
    list_display = (
        "exam_name",
        "class_obj",
        "exam_date",
    )

    list_filter = (
        "class_obj",
        "exam_date",
    )

    search_fields = (
        "exam_name",
    )

    autocomplete_fields = ("class_obj",)

    # ------------------------------------------------------------------
    # Same pattern as StudentPaymentAdmin: same admin URLs, custom
    # light-themed templates and views instead of Django's defaults.
    # Trade-off: ExamResultInline no longer shows on this page — results
    # are managed from the exam's own detail page instead.
    # ------------------------------------------------------------------
    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        exams = (
            Exam.objects.select_related("class_obj")
            .annotate(result_count=Count("examresult"))
            .order_by("-exam_date")
        )

        query = request.GET.get("q", "").strip()
        if query:
            exams = exams.filter(exam_name__icontains=query)

        active_class_id = request.GET.get("class", "").strip()
        if active_class_id:
            exams = exams.filter(class_obj_id=active_class_id)

        total_count = exams.count()
        paginator = Paginator(exams, 15)
        page_obj = paginator.get_page(request.GET.get("page"))

        preserved_params = request.GET.copy()
        preserved_params.pop("page", None)
        preserved_querystring = preserved_params.urlencode()

        context = dict(
            self.admin_site.each_context(request),
            exams=page_obj,
            page_obj=page_obj,
            page_range=_pagination_range(page_obj.number, paginator.num_pages),
            preserved_querystring=preserved_querystring,
            total_count=total_count,
            classes=Class.objects.order_by("-academic_year", "class_name"),
            active_class_id=active_class_id,
            search_query=query,
            active_section="academics",
            active_page="exam_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/exam/list.html", context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_exam_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._exam_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Exam, pk=object_id)
        return self._exam_form_view(request, instance=instance)

    def _exam_form_view(self, request, instance):
        if request.method == "POST":
            form = ExamForm(request.POST, instance=instance)
            if form.is_valid():
                exam = form.save()
                self.message_user(
                    request,
                    f"Exam {'updated' if instance else 'created'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_exam_view", exam.pk)
        else:
            form = ExamForm(instance=instance)

        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="academics",
            active_page="exam_edit" if instance else "exam_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/exam/form.html", context)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Exam.objects.select_related("class_obj"), pk=object_id)
        results = (
            ExamResult.objects.filter(exam=instance)
            .select_related("student", "subject")
            .order_by("student__full_name", "subject__subject_name")
        )
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            results=results,
            active_section="academics",
            active_page="exam_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/exam/detail.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(Exam, pk=object_id)
            obj.delete()
            self.message_user(request, "Exam deleted.", level=messages.SUCCESS)
        return redirect("admin:core_exam_changelist")


@admin.register(ExamResult)
class ExamResultAdmin(admin.ModelAdmin):
    list_display = (
        "student",
        "exam",
        "subject",
        "marks",
        "grade",
    )

    list_filter = (
        "exam",
        "subject",
        "grade",
    )

    search_fields = (
        "student__full_name",
        "student__student_code",
    )

    autocomplete_fields = ("student", "exam", "subject")

    # ------------------------------------------------------------------
    # No custom changelist here — results are always browsed from their
    # exam's detail page (admin:core_exam_view), not on their own, so
    # Django's default list page is left as a rarely-used fallback. Add,
    # change, and delete still get the dashboard-styled treatment and
    # redirect back to that exam page.
    # ------------------------------------------------------------------
    def get_urls(self):
        # Same pattern as StudentAdmin's bulk-upload/bulk-template pair,
        # but scoped to one exam at a time via ?exam=<id> — a mark sheet
        # only makes sense for a single exam's class and subject list.
        urls = super().get_urls()
        custom_urls = [
            path(
                "bulk-upload/",
                self.admin_site.admin_view(self.bulk_upload_view),
                name="core_examresult_bulk_upload",
            ),
            path(
                "bulk-template/",
                self.admin_site.admin_view(self.bulk_template_view),
                name="core_examresult_bulk_template",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._examresult_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(ExamResult, pk=object_id)
        return self._examresult_form_view(request, instance=instance)

    def _examresult_form_view(self, request, instance):
        # Arriving from a specific exam's "Add Result" link locks that
        # exam in as a hidden field, same reasoning as StudentPaymentForm's
        # lock_student — staff shouldn't have to pick it again.
        if instance is not None:
            locked_exam = instance.exam
        else:
            exam_id = request.GET.get("exam") or request.POST.get("exam")
            locked_exam = Exam.objects.filter(pk=exam_id).first() if exam_id else None

        if request.method == "POST":
            form = ExamResultForm(request.POST, instance=instance, lock_exam=locked_exam)
            if form.is_valid():
                result = form.save(commit=False)
                if locked_exam is not None:
                    result.exam = locked_exam
                result.save()
                self.message_user(
                    request,
                    f"Result {'updated' if instance else 'recorded'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_exam_view", result.exam_id)
        else:
            form = ExamResultForm(instance=instance, lock_exam=locked_exam)

        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            locked_exam=locked_exam,
            active_section="academics",
            active_page="examresult_edit" if instance else "examresult_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/examresult/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        obj = get_object_or_404(ExamResult, pk=object_id)
        exam_id = obj.exam_id
        if request.method == "POST":
            obj.delete()
            self.message_user(request, "Exam result deleted.", level=messages.SUCCESS)
        return redirect("admin:core_exam_view", exam_id)

    def bulk_template_view(self, request):
        if not self.has_add_permission(request):
            raise PermissionDenied
        exam_id = request.GET.get("exam")
        exam = get_object_or_404(Exam.objects.select_related("class_obj"), pk=exam_id) if exam_id else None
        if not exam:
            self.message_user(request, "Choose an exam first.", level=messages.ERROR)
            return redirect("admin:core_exam_changelist")

        subjects = list(
            Subject.objects.filter(class_obj=exam.class_obj).order_by("subject_name")
        )
        students = Student.objects.filter(class_obj=exam.class_obj).order_by("full_name")

        headers = BULK_EXAMRESULT_FIXED_HEADERS + [s.subject_name for s in subjects]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        sheet.append(headers)
        for student in students:
            sheet.append([student.student_code, student.full_name] + [None] * len(subjects))
        for col_idx in range(1, len(headers) + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        safe_name = re.sub(r"[^A-Za-z0-9]+", "_", exam.exam_name).strip("_") or "exam"
        response["Content-Disposition"] = f'attachment; filename="{safe_name}_results_template.xlsx"'
        return response

    def bulk_upload_view(self, request):
        if not self.has_add_permission(request):
            raise PermissionDenied
        exam_id = request.GET.get("exam") or request.POST.get("exam")
        exam = get_object_or_404(Exam.objects.select_related("class_obj"), pk=exam_id) if exam_id else None
        if not exam:
            self.message_user(request, "Choose an exam first.", level=messages.ERROR)
            return redirect("admin:core_exam_changelist")

        results = None
        if request.method == "POST":
            excel_file = request.FILES.get("excel_file")
            if not excel_file:
                self.message_user(request, "Please choose a file to upload.", level=messages.ERROR)
            else:
                results = self._process_bulk_upload(excel_file, exam)

        context = dict(
            self.admin_site.each_context(request),
            exam=exam,
            results=results,
            active_section="academics",
            active_page="examresult_bulk_upload",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/examresult/bulk_upload.html", context)

    def _process_bulk_upload(self, excel_file, exam):
        """
        Reads a mark-sheet workbook for one exam: Student ID, Full Name,
        then one column per subject in the exam's class. Each filled-in
        mark becomes an ExamResult — update_or_create, not create, so
        re-uploading a corrected sheet fixes existing marks rather than
        erroring on the duplicate exam+student+subject.

        Returns a dict with saved_count and errors (list of "Row N: ..."
        strings, plus any unrecognized subject columns).
        """
        try:
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception:
            return {"saved_count": 0, "errors": [
                "Couldn't read that file — make sure it's a valid .xlsx file, "
                "ideally the downloaded template with marks filled in."
            ]}

        sheet = workbook.active
        header_row = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
        if header_row[:2] != BULK_EXAMRESULT_FIXED_HEADERS:
            return {"saved_count": 0, "errors": [
                "The column headers don't match the template. Please download "
                "the template below and use it as-is."
            ]}

        subject_names = [h for h in header_row[2:] if h]
        subjects_by_name = {
            s.subject_name.strip().lower(): s
            for s in Subject.objects.filter(class_obj=exam.class_obj)
        }
        resolved_subjects = []
        unknown_subjects = []
        for name in header_row[2:]:
            subj = subjects_by_name.get(name.strip().lower()) if name else None
            resolved_subjects.append(subj)
            if name and not subj:
                unknown_subjects.append(name)

        errors = []
        if unknown_subjects:
            errors.append(
                "Unrecognized subject column(s), skipped: "
                + ", ".join(sorted(set(unknown_subjects)))
            )

        saved_count = 0
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v in (None, "") for v in row):
                continue  # skip fully blank rows

            student_code = str(row[0]).strip() if len(row) > 0 and row[0] not in (None, "") else ""
            if not student_code:
                continue
            student = Student.objects.filter(
                student_code__iexact=student_code, class_obj=exam.class_obj
            ).first()
            if not student:
                errors.append(
                    f"Row {row_num}: no student '{student_code}' in "
                    f"{exam.class_obj.class_name} — row skipped."
                )
                continue

            for col_idx, subject in enumerate(resolved_subjects, start=2):
                if subject is None or col_idx >= len(row):
                    continue
                raw_marks = row[col_idx]
                if raw_marks in (None, ""):
                    continue
                try:
                    marks = Decimal(str(raw_marks))
                except (InvalidOperation, ValueError):
                    errors.append(
                        f"Row {row_num} ({student.full_name}), {subject.subject_name}: "
                        f"'{raw_marks}' isn't a number — skipped."
                    )
                    continue
                if marks < 0:
                    errors.append(
                        f"Row {row_num} ({student.full_name}), {subject.subject_name}: "
                        f"marks can't be negative — skipped."
                    )
                    continue

                ExamResult.objects.update_or_create(
                    exam=exam, student=student, subject=subject,
                    defaults={"marks": marks},
                )
                saved_count += 1

        return {"saved_count": saved_count, "errors": errors}


@admin.register(StudentPayment)
class StudentPaymentAdmin(admin.ModelAdmin):
    list_display = (
        "student",
        "payment_type",
        "amount",
        "payment_date",
        "payment_month",
        "receipt_link",
    )

    list_filter = (
        "payment_type",
        "payment_date",
    )

    date_hierarchy = "payment_date"

    search_fields = (
        "student__full_name",
        "student__student_code",
    )

    autocomplete_fields = ("student",)

    # ------------------------------------------------------------------
    # Receipt link column on the changelist
    # ------------------------------------------------------------------
    def receipt_link(self, obj):
        url = reverse("admin:core_studentpayment_receipt", args=[obj.pk])
        return format_html('<a href="{}" target="_blank">View receipt</a>', url)
    receipt_link.short_description = "Receipt"

    # ------------------------------------------------------------------
    # Same pattern as StudentAdmin/ClassAdmin: same admin URLs, custom
    # light-themed templates and views instead of Django's defaults.
    # ------------------------------------------------------------------
    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        payments = StudentPayment.objects.select_related("student").order_by("-payment_date")

        query = request.GET.get("q", "").strip()
        if query:
            payments = payments.filter(
                Q(student__full_name__icontains=query)
                | Q(student__student_code__icontains=query)
            )

        active_payment_type = request.GET.get("payment_type", "").strip()
        if active_payment_type:
            payments = payments.filter(payment_type=active_payment_type)

        def _parse_iso_date(value):
            try:
                return datetime.strptime(value, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                return None

        date_from = request.GET.get("from", "").strip()
        parsed_from = _parse_iso_date(date_from) if date_from else None
        if parsed_from:
            payments = payments.filter(payment_date__gte=parsed_from)

        date_to = request.GET.get("to", "").strip()
        parsed_to = _parse_iso_date(date_to) if date_to else None
        if parsed_to:
            payments = payments.filter(payment_date__lte=parsed_to)

        total_count = payments.count()
        total_amount = payments.aggregate(total=Sum("amount"))["total"] or 0

        paginator = Paginator(payments, 10)
        page_obj = paginator.get_page(request.GET.get("page"))

        preserved_params = request.GET.copy()
        preserved_params.pop("page", None)
        preserved_querystring = preserved_params.urlencode()

        context = dict(
            self.admin_site.each_context(request),
            payments=page_obj,
            page_obj=page_obj,
            page_range=_pagination_range(page_obj.number, paginator.num_pages),
            preserved_querystring=preserved_querystring,
            total_count=total_count,
            total_amount=total_amount,
            payment_types=StudentPayment._meta.get_field("payment_type").choices,
            active_payment_type=active_payment_type,
            search_query=query,
            date_from=date_from,
            date_to=date_to,
            active_section="finance",
            active_page="studentpayment_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/studentpayment/list.html", context)

    def get_urls(self):
        # Adds a read-only "view" page alongside the existing add/change/
        # delete URLs, at admin/core/studentpayment/<id>/view/,
        # reverse()-able as admin:core_studentpayment_view — plus the
        # monthly fee status report and printable receipt.
        urls = super().get_urls()
        custom_urls = [
            path(
                "monthly-status/",
                self.admin_site.admin_view(self.monthly_status_view),
                name="core_studentpayment_monthly_status",
            ),
            path(
                "<int:payment_id>/receipt/",
                self.admin_site.admin_view(self.receipt_view),
                name="core_studentpayment_receipt",
            ),
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_studentpayment_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._studentpayment_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(StudentPayment, pk=object_id)
        return self._studentpayment_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(
            StudentPayment.objects.select_related(
                "student", "student__class_obj", "student__batch"
            ),
            pk=object_id,
        )
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            active_section="finance",
            active_page="studentpayment_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/studentpayment/detail.html", context)

    def _studentpayment_form_view(self, request, instance):
        # Arriving from a specific student's "Record a payment" link (or
        # the Monthly Fee Status report) passes ?student=<id> — lock that
        # student in as a hidden field instead of making staff pick them
        # again from the dropdown. Editing an existing payment locks it
        # to that payment's student the same way.
        if instance is not None:
            locked_student = instance.student
        else:
            student_id = request.GET.get("student") or request.POST.get("student")
            locked_student = (
                Student.objects.filter(pk=student_id).first() if student_id else None
            )

        if request.method == "POST":
            form = StudentPaymentForm(
                request.POST, instance=instance, lock_student=locked_student
            )
            if form.is_valid():
                payment = form.save(commit=False)
                if locked_student is not None:
                    payment.student = locked_student
                payment.save()
                self.message_user(
                    request,
                    f"Payment {'updated' if instance else 'recorded'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_studentpayment_changelist")
        else:
            initial = {}
            if instance is None:
                for key in ("payment_type", "payment_month", "payment_date"):
                    value = request.GET.get(key)
                    if value:
                        initial[key] = value
            form = StudentPaymentForm(
                instance=instance, initial=initial, lock_student=locked_student
            )

        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            locked_student=locked_student,
            active_section="finance",
            active_page="studentpayment_edit" if instance else "studentpayment_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/studentpayment/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(StudentPayment, pk=object_id)
            obj.delete()
            self.message_user(request, "Payment deleted.", level=messages.SUCCESS)
        return redirect("admin:core_studentpayment_changelist")

    def monthly_status_view(self, request):
        """
        Shows every active student for a given month and whether their
        Monthly Fee is Paid or Due. Admission Fee and Exam Fee are
        one-time charges, not recurring, so this report is scoped to
        Monthly Fee only — that's the fee type "due/paid per month"
        actually applies to.
        """
        month_param = request.GET.get("month")
        if month_param:
            try:
                target_month = datetime.strptime(month_param, "%Y-%m").date().replace(day=1)
            except ValueError:
                target_month = timezone.localdate().replace(day=1)
        else:
            target_month = timezone.localdate().replace(day=1)

        prev_month = (target_month - timedelta(days=1)).replace(day=1)
        next_month_probe = target_month.replace(day=28) + timedelta(days=4)
        next_month = next_month_probe.replace(day=1)

        students = (
            Student.objects.filter(status="Active")
            .select_related("class_obj", "batch")
            .order_by("full_name")
        )

        paid_qs = (
            StudentPayment.objects.filter(
                payment_type="Monthly Fee",
                payment_month__year=target_month.year,
                payment_month__month=target_month.month,
            )
            .values("student_id")
            .annotate(total=Sum("amount"))
        )
        latest_payment_by_student = {}
        for p in StudentPayment.objects.filter(
            payment_type="Monthly Fee",
            payment_month__year=target_month.year,
            payment_month__month=target_month.month,
        ).order_by("payment_date"):
            latest_payment_by_student[p.student_id] = p

        paid_totals = {p["student_id"]: p["total"] for p in paid_qs}

        rows = []
        paid_count = 0
        due_count = 0
        for student in students:
            amount_paid = paid_totals.get(student.id)
            payment_obj = latest_payment_by_student.get(student.id)
            is_paid = amount_paid is not None
            if is_paid:
                paid_count += 1
            else:
                due_count += 1

            add_url = (
                reverse("admin:core_studentpayment_add")
                + "?"
                + urlencode({
                    "student": student.id,
                    "payment_type": "Monthly Fee",
                    "payment_month": target_month.isoformat(),
                    "payment_date": timezone.localdate().isoformat(),
                })
            )

            rows.append({
                "student": student,
                "is_paid": is_paid,
                "amount_paid": amount_paid,
                "receipt_url": (
                    reverse("admin:core_studentpayment_receipt", args=[payment_obj.pk])
                    if payment_obj else None
                ),
                "record_payment_url": add_url,
            })

        context = dict(
            self.admin_site.each_context(request),
            title="Monthly Fee Status",
            rows=rows,
            target_month=target_month,
            target_month_label=target_month.strftime("%B %Y"),
            prev_month=prev_month.strftime("%Y-%m"),
            prev_month_label=prev_month.strftime("%b %Y"),
            next_month=next_month.strftime("%Y-%m"),
            next_month_label=next_month.strftime("%b %Y"),
            paid_count=paid_count,
            due_count=due_count,
            total_students=len(rows),
            active_section="finance",
            active_page="studentpayment_monthly_status",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/monthly_fee_status.html", context)

    def receipt_view(self, request, payment_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        payment = get_object_or_404(
            StudentPayment.objects.select_related(
                "student", "student__class_obj", "student__batch"
            ),
            pk=payment_id,
        )
        context = {
            "payment": payment,
            "academy_name": "Pi - π Academy",
            "generated_at": timezone.localtime(),
        }
        return TemplateResponse(request, "admin/core/payment_receipt.html", context)


@admin.register(TeacherSalary)
class TeacherSalaryAdmin(admin.ModelAdmin):
    list_display = (
        "teacher",
        "salary_month",
        "amount",
        "payment_date",
        "status",
        "receipt_link",
    )

    list_filter = (
        "status",
        "salary_month",
        "payment_date",
    )

    date_hierarchy = "salary_month"

    search_fields = (
        "teacher__full_name",
    )

    autocomplete_fields = ("teacher",)

    # ------------------------------------------------------------------
    # Receipt link column on the changelist
    # ------------------------------------------------------------------
    def receipt_link(self, obj):
        if obj.status != "Paid":
            return "—"
        url = reverse("admin:core_teachersalary_receipt", args=[obj.pk])
        return format_html('<a href="{}" target="_blank">View receipt</a>', url)
    receipt_link.short_description = "Receipt"

    # ------------------------------------------------------------------
    # Same pattern as StudentPaymentAdmin: same admin URLs, custom
    # light-themed templates and views instead of Django's defaults.
    # ------------------------------------------------------------------
    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        salaries = TeacherSalary.objects.select_related("teacher").order_by("-salary_month")

        query = request.GET.get("q", "").strip()
        if query:
            salaries = salaries.filter(teacher__full_name__icontains=query)

        active_status = request.GET.get("status", "").strip()
        if active_status:
            salaries = salaries.filter(status=active_status)

        total_count = salaries.count()
        total_amount = salaries.aggregate(total=Sum("amount"))["total"] or 0

        paginator = Paginator(salaries, 15)
        page_obj = paginator.get_page(request.GET.get("page"))

        preserved_params = request.GET.copy()
        preserved_params.pop("page", None)
        preserved_querystring = preserved_params.urlencode()

        context = dict(
            self.admin_site.each_context(request),
            salaries=page_obj,
            page_obj=page_obj,
            page_range=_pagination_range(page_obj.number, paginator.num_pages),
            preserved_querystring=preserved_querystring,
            total_count=total_count,
            total_amount=total_amount,
            status_choices=TeacherSalary._meta.get_field("status").choices,
            active_status=active_status,
            search_query=query,
            active_section="finance",
            active_page="teachersalary_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/teachersalary/list.html", context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "monthly-status/",
                self.admin_site.admin_view(self.salary_status_view),
                name="core_teachersalary_monthly_status",
            ),
            path(
                "<int:salary_id>/receipt/",
                self.admin_site.admin_view(self.receipt_view),
                name="core_teachersalary_receipt",
            ),
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_teachersalary_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._teachersalary_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(TeacherSalary, pk=object_id)
        return self._teachersalary_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(
            TeacherSalary.objects.select_related("teacher"), pk=object_id
        )
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            active_section="finance",
            active_page="teachersalary_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/teachersalary/detail.html", context)

    def _teachersalary_form_view(self, request, instance):
        # Arriving from a specific faculty member's "Record Salary" link
        # (or the Monthly Salary Status report) passes ?teacher=<id> —
        # lock that teacher in as a hidden field, same reasoning as
        # StudentPaymentForm's lock_student. Editing an existing record
        # locks it to that record's teacher the same way.
        if instance is not None:
            locked_teacher = instance.teacher
        else:
            teacher_id = request.GET.get("teacher") or request.POST.get("teacher")
            locked_teacher = (
                Faculty.objects.filter(pk=teacher_id).first() if teacher_id else None
            )

        if request.method == "POST":
            form = TeacherSalaryForm(
                request.POST, instance=instance, lock_teacher=locked_teacher
            )
            if form.is_valid():
                salary = form.save(commit=False)
                if locked_teacher is not None:
                    salary.teacher = locked_teacher
                salary.save()
                self.message_user(
                    request,
                    f"Salary record {'updated' if instance else 'saved'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_teachersalary_changelist")
        else:
            initial = {}
            if instance is None:
                salary_month = request.GET.get("salary_month")
                if salary_month:
                    initial["salary_month"] = salary_month
            form = TeacherSalaryForm(
                instance=instance, initial=initial, lock_teacher=locked_teacher
            )

        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            locked_teacher=locked_teacher,
            active_section="finance",
            active_page="teachersalary_edit" if instance else "teachersalary_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/teachersalary/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(TeacherSalary, pk=object_id)
            obj.delete()
            self.message_user(request, "Salary record deleted.", level=messages.SUCCESS)
        return redirect("admin:core_teachersalary_changelist")

    def salary_status_view(self, request):
        """
        Shows every faculty member for a given month, and whether their
        fixed ৳15,000 salary is Paid, Pending (a record exists but hasn't
        been marked paid), or hasn't been recorded at all yet.
        """
        month_param = request.GET.get("month")
        if month_param:
            try:
                target_month = datetime.strptime(month_param, "%Y-%m").date().replace(day=1)
            except ValueError:
                target_month = timezone.localdate().replace(day=1)
        else:
            target_month = timezone.localdate().replace(day=1)

        prev_month = (target_month - timedelta(days=1)).replace(day=1)
        next_month_probe = target_month.replace(day=28) + timedelta(days=4)
        next_month = next_month_probe.replace(day=1)

        teachers = Faculty.objects.all().order_by("full_name")

        salaries_by_teacher = {
            s.teacher_id: s
            for s in TeacherSalary.objects.filter(
                salary_month__year=target_month.year,
                salary_month__month=target_month.month,
            )
        }

        rows = []
        paid_count = 0
        pending_count = 0
        not_recorded_count = 0
        for teacher in teachers:
            salary = salaries_by_teacher.get(teacher.id)
            if salary is None:
                row_status = "not_recorded"
                not_recorded_count += 1
                action_url = (
                    reverse("admin:core_teachersalary_add")
                    + "?"
                    + urlencode({
                        "teacher": teacher.id,
                        "salary_month": target_month.strftime("%Y-%m"),
                    })
                )
            elif salary.status == "Paid":
                row_status = "paid"
                paid_count += 1
                action_url = reverse("admin:core_teachersalary_change", args=[salary.pk])
            else:
                row_status = "pending"
                pending_count += 1
                action_url = reverse("admin:core_teachersalary_change", args=[salary.pk])

            rows.append({
                "teacher": teacher,
                "salary": salary,
                "status": row_status,
                "action_url": action_url,
                "receipt_url": (
                    reverse("admin:core_teachersalary_receipt", args=[salary.pk])
                    if salary and salary.status == "Paid" else None
                ),
            })

        context = dict(
            self.admin_site.each_context(request),
            title="Monthly Salary Status",
            rows=rows,
            target_month=target_month,
            target_month_label=target_month.strftime("%B %Y"),
            prev_month=prev_month.strftime("%Y-%m"),
            prev_month_label=prev_month.strftime("%b %Y"),
            next_month=next_month.strftime("%Y-%m"),
            next_month_label=next_month.strftime("%b %Y"),
            paid_count=paid_count,
            pending_count=pending_count,
            not_recorded_count=not_recorded_count,
            total_teachers=len(rows),
            fixed_salary_amount=15000,
            active_section="finance",
            active_page="teachersalary_monthly_status",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/monthly_salary_status.html", context)

    def receipt_view(self, request, salary_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        salary = get_object_or_404(
            TeacherSalary.objects.select_related("teacher"), pk=salary_id
        )
        context = {
            "salary": salary,
            "academy_name": "Pi - π Academy",
            "generated_at": timezone.localtime(),
        }
        return TemplateResponse(request, "admin/core/salary_receipt.html", context)


@admin.register(Achievement)
class AchievementAdmin(admin.ModelAdmin):
    search_fields = ("title", "description")

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        achievements = Achievement.objects.all().order_by("-achievement_date", "-created_at")
        context = dict(
            self.admin_site.each_context(request),
            achievements=achievements,
            active_section="content",
            active_page="achievement_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/achievement/list.html", context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_achievement_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._achievement_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Achievement, pk=object_id)
        return self._achievement_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Achievement, pk=object_id)
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            active_section="content",
            active_page="achievement_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/achievement/detail.html", context)

    def _achievement_form_view(self, request, instance):
        if request.method == "POST":
            form = AchievementForm(request.POST, request.FILES, instance=instance)
            if form.is_valid():
                form.save()
                self.message_user(
                    request,
                    f"Achievement {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_achievement_changelist")
        else:
            form = AchievementForm(instance=instance)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="content",
            active_page="achievement_edit" if instance else "achievement_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/achievement/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(Achievement, pk=object_id)
            obj.delete()
            self.message_user(request, "Achievement deleted.", level=messages.SUCCESS)
        return redirect("admin:core_achievement_changelist")


@admin.register(Gallery)
class GalleryAdmin(admin.ModelAdmin):
    search_fields = ("title", "description", "category")

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        items = Gallery.objects.all().order_by("-created_at")
        context = dict(
            self.admin_site.each_context(request),
            items=items,
            active_section="content",
            active_page="gallery_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/gallery/list.html", context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_gallery_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._gallery_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Gallery, pk=object_id)
        return self._gallery_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Gallery, pk=object_id)
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            active_section="content",
            active_page="gallery_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/gallery/detail.html", context)

    def _gallery_form_view(self, request, instance):
        if request.method == "POST":
            form = GalleryForm(request.POST, request.FILES, instance=instance)
            if form.is_valid():
                form.save()
                self.message_user(
                    request,
                    f"Gallery item {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_gallery_changelist")
        else:
            form = GalleryForm(instance=instance)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="content",
            active_page="gallery_edit" if instance else "gallery_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/gallery/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(Gallery, pk=object_id)
            obj.delete()
            self.message_user(request, "Gallery item deleted.", level=messages.SUCCESS)
        return redirect("admin:core_gallery_changelist")


@admin.register(Notice)
class NoticeAdmin(admin.ModelAdmin):
    search_fields = ("title", "description")

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        notices = Notice.objects.all().order_by("-published_date", "-created_at")
        context = dict(
            self.admin_site.each_context(request),
            notices=notices,
            active_section="content",
            active_page="notice_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/notice/list.html", context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_notice_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._notice_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Notice, pk=object_id)
        return self._notice_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(Notice, pk=object_id)
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            active_section="content",
            active_page="notice_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/notice/detail.html", context)

    def _notice_form_view(self, request, instance):
        if request.method == "POST":
            form = NoticeForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                self.message_user(
                    request,
                    f"Notice {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_notice_changelist")
        else:
            form = NoticeForm(instance=instance)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="content",
            active_page="notice_edit" if instance else "notice_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/notice/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(Notice, pk=object_id)
            obj.delete()
            self.message_user(request, "Notice deleted.", level=messages.SUCCESS)
        return redirect("admin:core_notice_changelist")


@admin.register(AdmissionInfo)
class AdmissionInfoAdmin(admin.ModelAdmin):
    search_fields = ("title", "description")

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        info_items = AdmissionInfo.objects.all().order_by("-created_at")
        context = dict(
            self.admin_site.each_context(request),
            info_items=info_items,
            active_section="content",
            active_page="admissioninfo_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/admissioninfo/list.html", context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_admissioninfo_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._admissioninfo_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(AdmissionInfo, pk=object_id)
        return self._admissioninfo_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(AdmissionInfo, pk=object_id)
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            active_section="content",
            active_page="admissioninfo_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/admissioninfo/detail.html", context)

    def _admissioninfo_form_view(self, request, instance):
        if request.method == "POST":
            form = AdmissionInfoForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                self.message_user(
                    request,
                    f"Admission info {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_admissioninfo_changelist")
        else:
            form = AdmissionInfoForm(instance=instance)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="content",
            active_page="admissioninfo_edit" if instance else "admissioninfo_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/admissioninfo/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(AdmissionInfo, pk=object_id)
            obj.delete()
            self.message_user(request, "Admission info deleted.", level=messages.SUCCESS)
        return redirect("admin:core_admissioninfo_changelist")


@admin.register(AdmissionInquiry)
class AdmissionInquiryAdmin(admin.ModelAdmin):
    search_fields = ("student_name", "student_phone", "guardian_name", "guardian_phone")

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        inquiries = AdmissionInquiry.objects.select_related("class_obj").order_by("-created_at")
        active_status = request.GET.get("status")
        if active_status:
            inquiries = inquiries.filter(status=active_status)
        context = dict(
            self.admin_site.each_context(request),
            inquiries=inquiries,
            status_choices=AdmissionInquiry._meta.get_field("status").choices,
            active_status=active_status,
            active_section="inbox",
            active_page="admissioninquiry_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/admissioninquiry/list.html", context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_admissioninquiry_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._admissioninquiry_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(AdmissionInquiry, pk=object_id)
        return self._admissioninquiry_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(
            AdmissionInquiry.objects.select_related("class_obj"), pk=object_id
        )
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            active_section="inbox",
            active_page="admissioninquiry_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/admissioninquiry/detail.html", context)

    def _admissioninquiry_form_view(self, request, instance):
        if request.method == "POST":
            form = AdmissionInquiryAdminForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                self.message_user(
                    request,
                    f"Admission inquiry {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_admissioninquiry_changelist")
        else:
            form = AdmissionInquiryAdminForm(instance=instance)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="inbox",
            active_page="admissioninquiry_edit" if instance else "admissioninquiry_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/admissioninquiry/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(AdmissionInquiry, pk=object_id)
            obj.delete()
            self.message_user(request, "Admission inquiry deleted.", level=messages.SUCCESS)
        return redirect("admin:core_admissioninquiry_changelist")


@admin.register(ContactMessage)
class ContactMessageAdmin(admin.ModelAdmin):
    search_fields = ("name", "phone", "email", "message")

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_permission(request):
            raise PermissionDenied
        contact_messages = ContactMessage.objects.all().order_by("-created_at")
        active_status = request.GET.get("status")
        if active_status:
            contact_messages = contact_messages.filter(status=active_status)
        context = dict(
            self.admin_site.each_context(request),
            contact_messages=contact_messages,
            status_choices=ContactMessage._meta.get_field("status").choices,
            active_status=active_status,
            active_section="inbox",
            active_page="contactmessage_list",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/contactmessage/list.html", context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/view/",
                self.admin_site.admin_view(self.view_view),
                name="core_contactmessage_view",
            ),
        ]
        return custom_urls + urls

    def add_view(self, request, form_url="", extra_context=None):
        if not self.has_add_permission(request):
            raise PermissionDenied
        return self._contactmessage_form_view(request, instance=None)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if not self.has_change_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(ContactMessage, pk=object_id)
        return self._contactmessage_form_view(request, instance=instance)

    def view_view(self, request, object_id):
        if not self.has_view_permission(request):
            raise PermissionDenied
        instance = get_object_or_404(ContactMessage, pk=object_id)
        context = dict(
            self.admin_site.each_context(request),
            instance=instance,
            active_section="inbox",
            active_page="contactmessage_view",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/contactmessage/detail.html", context)

    def _contactmessage_form_view(self, request, instance):
        if request.method == "POST":
            form = ContactMessageAdminForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                self.message_user(
                    request,
                    f"Contact message {'updated' if instance else 'added'} successfully.",
                    level=messages.SUCCESS,
                )
                return redirect("admin:core_contactmessage_changelist")
        else:
            form = ContactMessageAdminForm(instance=instance)
        context = dict(
            self.admin_site.each_context(request),
            form=form,
            instance=instance,
            active_section="inbox",
            active_page="contactmessage_edit" if instance else "contactmessage_add",
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/core/contactmessage/form.html", context)

    def delete_view(self, request, object_id, extra_context=None):
        if not self.has_delete_permission(request):
            raise PermissionDenied
        if request.method == "POST":
            obj = get_object_or_404(ContactMessage, pk=object_id)
            obj.delete()
            self.message_user(request, "Contact message deleted.", level=messages.SUCCESS)
        return redirect("admin:core_contactmessage_changelist")


# ---------------------------------------------------------
# Dashboard — inject live stats + a full categorized management
# grid into the admin homepage context
# ---------------------------------------------------------

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.urls import reverse

_default_each_context = admin.site.each_context

# (display label, url-name key, queryset) for every manageable model,
# grouped into the categories shown on the dashboard.
_DASHBOARD_LAYOUT = [
    ("People", [
        ("Students", "core_student", lambda: Student.objects.count()),
        ("Faculty", "core_faculty", lambda: Faculty.objects.count()),
    ]),
    ("Academics", [
        ("Classes", "core_class", lambda: Class.objects.count()),
        ("Batches", "core_batch", lambda: Batch.objects.count()),
        ("Subjects", "core_subject", lambda: Subject.objects.count()),
        ("Class Schedules", "core_classschedule", lambda: ClassSchedule.objects.count()),
    ]),
    ("Records", [
        ("Attendance Records", "core_attendance", lambda: Attendance.objects.count()),
        ("Exams", "core_exam", lambda: Exam.objects.count()),
        ("Exam Results", "core_examresult", lambda: ExamResult.objects.count()),
    ]),
    ("Finance", [
        ("Student Payments", "core_studentpayment", lambda: StudentPayment.objects.count()),
        ("Teacher Salaries", "core_teachersalary", lambda: TeacherSalary.objects.count()),
    ]),
    ("Website Content", [
        ("Achievements", "core_achievement", lambda: Achievement.objects.count()),
        ("Gallery Items", "core_gallery", lambda: Gallery.objects.count()),
        ("Notices", "core_notice", lambda: Notice.objects.count()),
        ("Admission Infos", "core_admissioninfo", lambda: AdmissionInfo.objects.count()),
    ]),
    ("Inbox", [
        ("Admission Inquiries", "core_admissioninquiry", lambda: AdmissionInquiry.objects.filter(status="New").count()),
        ("Contact Messages", "core_contactmessage", lambda: ContactMessage.objects.filter(status="Unread").count()),
    ]),
    ("System", [
        ("Users", "auth_user", lambda: get_user_model().objects.count()),
        ("Groups", "auth_group", lambda: Group.objects.count()),
    ]),
]

# Inbox counts above are "needs attention" counts (new/unread), so flag those cards
_ALERT_SECTIONS = {"Inbox"}

_SECTION_ICONS = {
    "People": "icon-user",
    "Academics": "icon-book",
    "Records": "icon-clipboard",
    "Finance": "icon-wallet",
    "Website Content": "icon-image",
    "Inbox": "icon-mail",
    "System": "icon-gear",
}


def _build_dashboard_sections():
    sections = []
    for label, items in _DASHBOARD_LAYOUT:
        cards = []
        for name, url_key, count_fn in items:
            cards.append({
                "name": name,
                "count": count_fn(),
                "add_url_name": f"admin:{url_key}_add",
                "list_url_name": f"admin:{url_key}_changelist",
            })
        sections.append({
            "label": label,
            "cards": cards,
            "is_alert": label in _ALERT_SECTIONS,
            "icon": _SECTION_ICONS.get(label, "icon-book"),
            "total": sum(c["count"] for c in cards),
        })
    return sections


def _each_context_with_dashboard(request):
    context = _default_each_context(request)

    try:
        is_index = request.path == reverse("admin:index")
    except Exception:
        is_index = False

    if not is_index:
        return context

    today = timezone.localdate()
    month_start = today.replace(day=1)

    todays_attendance = Attendance.objects.filter(attendance_date=today)
    monthly_revenue = StudentPayment.objects.filter(
        payment_date__gte=month_start, payment_date__lte=today
    ).aggregate(total=Sum("amount"))["total"] or 0

    context["dashboard_stats"] = {
        "active_students": Student.objects.filter(status="Active").count(),
        "total_faculty": Faculty.objects.count(),
        "pending_inquiries": AdmissionInquiry.objects.filter(status="New").count(),
        "unread_messages": ContactMessage.objects.filter(status="Unread").count(),
        "today_present": todays_attendance.filter(status="Present").count(),
        "today_total": todays_attendance.count(),
        "monthly_revenue": monthly_revenue,
        "pending_salaries": TeacherSalary.objects.filter(status="Pending").count(),
    }
    context["dashboard_sections"] = _build_dashboard_sections()

    # Restores the audit-trail visibility that the default admin sidebar
    # used to provide (we removed {% block sidebar %} to fix the blank
    # right-column layout bug) — folded into the custom dashboard instead
    # of the default floating panel.
    context["dashboard_recent_actions"] = (
        LogEntry.objects.select_related("content_type", "user")
        .filter(user_id=request.user.pk)
        .order_by("-action_time")[:8]
    )

    hour = timezone.localtime().hour
    if hour < 12:
        greeting = "Good morning"
    elif hour < 17:
        greeting = "Good afternoon"
    else:
        greeting = "Good evening"
    context["dashboard_greeting"] = greeting
    context["dashboard_today"] = today.strftime("%A, %d %B %Y")

    return context


admin.site.each_context = _each_context_with_dashboard